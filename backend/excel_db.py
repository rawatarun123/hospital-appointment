# ============================================================
#  excel_db.py — Excel storage helpers
#  All data lives in .xlsx files in the project root.
# ============================================================

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# File paths
F_USERS   = os.path.join(BASE, "users.xlsx")
F_DOCTORS = os.path.join(BASE, "doctors.xlsx")
F_PATIENTS= os.path.join(BASE, "patients.xlsx")
F_APPTS   = os.path.join(BASE, "appointments.xlsx")

# Column headers
H_USERS   = ["ID","Username","Password","Role","Name","Email","Phone","Status","Created_At"]
H_DOCTORS = ["ID","User_ID","Name","Specialization","Start_Time","End_Time","Slot_Duration","Experience","Fee","Status","Created_At"]
H_PATIENTS= ["ID","User_ID","Name","Age","Phone","Email","Blood_Group","Medical_History","Status","Created_At"]
H_APPTS   = ["ID","Patient_ID","Patient_Name","Patient_Phone","Doctor_ID","Doctor_Name",
             "Specialty","Symptoms","Urgency","Appt_Date","Appt_Time",
             "Priority","Status","Notes","Created_At"]


def read(filepath):
    """Read Excel sheet → list of dicts."""
    wb   = openpyxl.load_workbook(filepath)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h) for h in rows[0]]
    return [
        dict(zip(headers, row))
        for row in rows[1:]
        if any(v is not None for v in row)
    ]


def write(filepath, headers, rows):
    """Write list of dicts to Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="1D4ED8")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([row.get(h) for h in headers])
    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(mx + 4, 45)
    wb.save(filepath)
    wb.close()


def next_id(rows):
    ids = [int(r.get("ID") or 0) for r in rows]
    return max(ids, default=0) + 1


def now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def setup_all():
    """Create all Excel files with seed data on first run."""
    import hashlib

    def sha(p): return hashlib.sha256(p.encode()).hexdigest()

    if not os.path.exists(F_USERS):
        write(F_USERS, H_USERS, [
            {"ID":1,"Username":"admin",   "Password":sha("admin123"),  "Role":"admin",   "Name":"Admin User",       "Email":"admin@haso.com",   "Phone":"0000000000","Status":"active","Created_At":now()},
            {"ID":2,"Username":"drpatel", "Password":sha("doc123"),    "Role":"doctor",  "Name":"Dr. Aisha Patel",  "Email":"patel@haso.com",   "Phone":"9876543210","Status":"active","Created_At":now()},
            {"ID":3,"Username":"drchen",  "Password":sha("doc123"),    "Role":"doctor",  "Name":"Dr. Marcus Chen",  "Email":"chen@haso.com",    "Phone":"9876543211","Status":"active","Created_At":now()},
            {"ID":4,"Username":"alice",   "Password":sha("pat123"),    "Role":"patient", "Name":"Alice Johnson",    "Email":"alice@mail.com",   "Phone":"9876500001","Status":"active","Created_At":now()},
            {"ID":5,"Username":"bob",     "Password":sha("pat123"),    "Role":"patient", "Name":"Bob Williams",     "Email":"bob@mail.com",     "Phone":"9876500002","Status":"active","Created_At":now()},
        ])
        print("  ✓ users.xlsx")

    if not os.path.exists(F_DOCTORS):
        write(F_DOCTORS, H_DOCTORS, [
            {"ID":1,"User_ID":2,"Name":"Dr. Aisha Patel",  "Specialization":"Cardiology",  "Start_Time":"09:00","End_Time":"17:00","Slot_Duration":30,"Experience":12,"Fee":500,"Status":"available","Created_At":now()},
            {"ID":2,"User_ID":3,"Name":"Dr. Marcus Chen",  "Specialization":"Neurology",   "Start_Time":"10:00","End_Time":"16:00","Slot_Duration":45,"Experience":8, "Fee":600,"Status":"available","Created_At":now()},
            {"ID":3,"User_ID":None,"Name":"Dr. Sofia Reyes","Specialization":"Orthopedics","Start_Time":"09:00","End_Time":"13:00","Slot_Duration":30,"Experience":10,"Fee":550,"Status":"available","Created_At":now()},
            {"ID":4,"User_ID":None,"Name":"Dr. James Okafor","Specialization":"General",   "Start_Time":"08:00","End_Time":"18:00","Slot_Duration":20,"Experience":15,"Fee":300,"Status":"available","Created_At":now()},
            {"ID":5,"User_ID":None,"Name":"Dr. Priya Nair", "Specialization":"Pediatrics", "Start_Time":"09:00","End_Time":"15:00","Slot_Duration":30,"Experience":7, "Fee":400,"Status":"available","Created_At":now()},
            {"ID":6,"User_ID":None,"Name":"Dr. Ravi Sharma","Specialization":"Dermatology","Start_Time":"11:00","End_Time":"17:00","Slot_Duration":20,"Experience":9, "Fee":450,"Status":"available","Created_At":now()},
        ])
        print("  ✓ doctors.xlsx")

    if not os.path.exists(F_PATIENTS):
        write(F_PATIENTS, H_PATIENTS, [
            {"ID":1,"User_ID":4,"Name":"Alice Johnson", "Age":34,"Phone":"9876500001","Email":"alice@mail.com","Blood_Group":"O+", "Medical_History":"None","Status":"active","Created_At":now()},
            {"ID":2,"User_ID":5,"Name":"Bob Williams",  "Age":52,"Phone":"9876500002","Email":"bob@mail.com",  "Blood_Group":"A+", "Medical_History":"Hypertension","Status":"active","Created_At":now()},
        ])
        print("  ✓ patients.xlsx")

    if not os.path.exists(F_APPTS):
        write(F_APPTS, H_APPTS, [])
        print("  ✓ appointments.xlsx")
